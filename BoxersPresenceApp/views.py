# ===== BATTERY TESTS =====
import io
from collections import defaultdict
from datetime import datetime, timedelta
from datetime import date, time as dt_time
from datetime import date as Date
from decimal import InvalidOperation
import openpyxl
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.db import transaction, IntegrityError
from django.db.models import Max, ProtectedError, Min, Avg, Count
from django.db.models.functions import TruncDate
from django.forms import formset_factory
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse, HttpResponseForbidden, Http404, HttpRequest
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.views.generic import UpdateView, FormView, ListView, DetailView, CreateView
from django.views import View
from django.urls import reverse_lazy, reverse, get_resolver
from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl.utils import get_column_letter
from rest_framework.exceptions import PermissionDenied
from .models import Boxer, BatteryTest, Attendance, HeartRate, Weight, ParentProfile, ClassTemplate, \
    Enrollment, Gym, BoxerComment
from .forms import BatteryTestForm, BoxerAndTestSelectForm, BoxerForm, HeartRateQuickForm, \
    WeightQuickForm, ParentSignupForm, GymForm, TestResultForm, EnrollBoxerForm, ClassCreateForm, UnenrollForm, \
    BulkBoxerForm, BoxerCommentForm
from .models import TestResult
from .utils import user_gym, resolve_boxer, expand_rrule

from decimal import Decimal
from django.db.models import Exists, OuterRef, Q
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin


# --- phase helpers (ensure these exist in the file once) ---
PHASE_SYNONYMS = {
    "prep":  {"prep", "pre", "preparation", "Pre", "PRE", "Preparation"},
    "build": {"build", "mid", "Mid", "BUILD", "Build"},
    "peak":  {"peak", "before", "Before", "before tournament", "Before Tournament", "PEAK", "Peak"},
}
def normalize_phase(raw: str) -> str:
    raw = (raw or "").strip().lower()
    for canon, syns in PHASE_SYNONYMS.items():
        if raw in {s.lower() for s in syns}:
            return canon
    return "prep"
def phase_family(slug: str):
    canon = normalize_phase(slug)
    return PHASE_SYNONYMS.get(canon, {canon})

def is_parent_user(user):
    try:
        return bool(user.parent_profile)
    except ParentProfile.DoesNotExist:
        return False


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)

        # Who can you see?
        if self.request.user.is_superuser:
            boxers_qs = Boxer.objects.all()
        else:
            if gym:
                boxers_qs = Boxer.objects.filter(
                    Q(gym=gym) | Q(shared_with_gyms=gym) | Q(coaches=self.request.user)
                ).distinct()
            else:
                boxers_qs = Boxer.objects.filter(coaches=self.request.user).distinct()

        boxers_qs = boxers_qs.order_by("name")

        # Build the HR quick form and restrict its boxer queryset to what you can see
        hr_form = HeartRateQuickForm()
        if "boxer" in hr_form.fields and hasattr(hr_form.fields["boxer"], "queryset"):
            hr_form.fields["boxer"].queryset = boxers_qs

        ctx.update({
            "boxers": boxers_qs,
            "hr_form": hr_form,
        })
        return ctx

class RegisterView(FormView):
    template_name = 'login/register.html'
    form_class = UserCreationForm
    success_url = reverse_lazy('home')

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        return super().form_valid(form)

def _user_gym(user):
    """Return the user's gym; create/get a default if missing."""
    try:
        return user.coach_profile.gym
    except Exception:
        gym, _ = Gym.objects.get_or_create(name="Default Gym", defaults={"timezone": "Europe/Brussels"})
        return gym

class BoxerListView(LoginRequiredMixin, ListView):
    template_name = 'boxers/boxer_list.html'
    context_object_name = 'boxers'
    paginate_by = 25

    def get_queryset(self):
        gym = user_gym(self.request)
        q = (self.request.GET.get("q") or "").strip()
        class_id = self.request.GET.get("class_id")

        qs = Boxer.objects.filter(gym=gym)

        if class_id:
            qs = qs.filter(enrollments__template_id=class_id)

        if q:
            for term in q.split():
                qs = qs.filter(
                    Q(name__icontains=term) |
                    Q(parent_name__icontains=term)
                )
        return qs.order_by("first_name", "last_name", "name")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)

        ctx["q"] = (self.request.GET.get("q") or "").strip()
        ctx["class_id"] = self.request.GET.get("class_id") or ""
        ctx["result_count"] = self.get_queryset().count()
        ctx["form"] = BoxerForm()
        ctx["classes"] = ClassTemplate.objects.filter(gym=gym).order_by("title")

        return ctx

class TestsListView(LoginRequiredMixin, TemplateView):
    template_name = "tests/tests_list.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["form"] = BatteryTestForm()
        ctx["tests"] = BatteryTest.objects.all().order_by("display_order", "name")
        return ctx

    def post(self, request, *args, **kwargs):
        form = BatteryTestForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("tests_list")
        # fall back: re-render with errors
        ctx = self.get_context_data()
        ctx["form"] = form
        return self.render_to_response(ctx)


class TestUpdateView(LoginRequiredMixin, UpdateView):
    model = BatteryTest
    form_class = BatteryTestForm
    template_name = "tests/test_edit.html"
    pk_url_kwarg = "pk"
    success_url = reverse_lazy("tests_list")

    # Don’t scope/hide tests here. Show all so the link always works.
    def get_queryset(self):
        return BatteryTest.objects.all()

    # Optional: friendlier than a plain 404
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not obj:
            messages.error(self.request, "That test doesn’t exist (maybe it was deleted).")
            raise Http404
        return obj


# views.py
class BatteryTestDeleteView(View):
    success_url = reverse_lazy("tests_list")

    def post(self, request, pk, *args, **kwargs):
        test = get_object_or_404(BatteryTest, pk=pk)
        try:
            with transaction.atomic():
                test.delete()
        except ProtectedError:
            # If your FKs are PROTECT/RESTRICT and you want to cascade in code:
            with transaction.atomic():
                TestResult.objects.filter(test=test).delete()
                test.delete()
        except IntegrityError:
            messages.error(request, "Could not delete this test due to database constraints.")
            return redirect(self.success_url)

        messages.success(request, f'“{test.name}” deleted.')
        return redirect(self.success_url)

    def get(self, request, pk, *args, **kwargs):
        # No deletes via GET; just bounce back
        return redirect(self.success_url)


def post(self, request, *args, **kwargs):
    # Save a single (boxer, test, phase) result from the detail form
    phase    = request.POST.get('phase') or TestResult.PHASE_PRE
    boxer_id = request.POST.get('boxer_id')
    test_id  = request.POST.get('test_id')

    try:
        boxer = get_object_or_404(Boxer, id=int(boxer_id), coach=request.user)
        test  = get_object_or_404(BatteryTest, id=int(test_id), coach=request.user)
    except (TypeError, ValueError):
        messages.error(request, "Please select a boxer and a test.")
        return redirect(f"{reverse_lazy('tests_results')}?phase={phase}")

    def parse_val(name):
        raw = (request.POST.get(name) or '').strip()
        if raw == '':
            return None
        try:
            return float(raw)
        except ValueError:
            raise ValidationError(f"{name} must be a number or blank.")

    try:
        v1 = parse_val('value1')
        v2 = parse_val('value2')
        v3 = parse_val('value3')
    except ValidationError as e:
        messages.error(request, str(e))
        return redirect(f"{reverse_lazy('tests_results')}?boxer={boxer.id}&test={test.id}&phase={phase}")

    notes = (request.POST.get('notes') or '').strip()

    obj, _ = TestResult.objects.get_or_create(boxer=boxer, test=test, phase=phase)
    obj.value1, obj.value2, obj.value3, obj.notes = v1, v2, v3, notes
    obj.save()

    messages.success(request, "Result saved.")
    return redirect(f"{reverse_lazy('tests_results')}?boxer={boxer.id}&test={test.id}&phase={phase}")


class ResultsCellSaveView(LoginRequiredMixin, View):
    """Save a single (boxer,test) cell for the selected phase."""
    def post(self, request, *args, **kwargs):
        phase = request.GET.get('phase') or request.POST.get('phase') or TestResult.PHASE_PRE
        boxer_id = request.GET.get('b') or request.POST.get('boxer_id')
        test_id  = request.GET.get('t') or request.POST.get('test_id')

        # If params are missing/invalid -> just go back to the Tests list (NO popup)
        try:
            boxer_id, test_id = int(boxer_id), int(test_id)
        except (TypeError, ValueError):
            return redirect("tests_list")

        # Do NOT filter by non-existent fields like coach=...
        # If you want to scope boxers to the current coach, use coaches=request.user (M2M).
        boxer = get_object_or_404(Boxer, id=boxer_id)  # or: id=boxer_id, coaches=request.user
        test  = get_object_or_404(BatteryTest, id=test_id)

        prefix = f"r-{boxer_id}-{test_id}-"
        v1 = request.POST.get(prefix + "value1") or None
        v2 = request.POST.get(prefix + "value2") or None
        v3 = request.POST.get(prefix + "value3") or None
        notes = (request.POST.get(prefix + "notes") or "").strip()

        obj, _ = TestResult.objects.get_or_create(boxer=boxer, test=test, phase=phase)
        obj.value1, obj.value2, obj.value3 = v1, v2, v3
        obj.notes = notes
        obj.save()

        # Always return to the Tests list (not the removed results matrix)
        return redirect("tests_list")



# ===== ONE BOXER (not phase-specific; keep simple) =====

# ===== SUMMARY (not phase-specific) =====

# ===== MANAGE BOXERS (boxer + test, show best per phase + improvement) =====
class BoxerPerformanceView(LoginRequiredMixin, FormView):
    template_name = 'tests/tests_manage_boxers.html'
    form_class = BoxerAndTestSelectForm
    success_url = reverse_lazy('tests_manage_boxers')

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['user'] = self.request.user
        return kw

    def form_valid(self, form):
        boxer = form.cleaned_data['boxer']
        test  = form.cleaned_data['test']

        def best_for(phase):
            tr = TestResult.objects.filter(boxer=boxer, test=test, phase=phase).first()
            if not tr:
                return None
            vals = [v for v in (tr.value1, tr.value2, tr.value3) if v is not None]
            if not vals:
                return None
            unit = (test.unit or '').strip().lower()
            return min(vals) if unit == 's' else max(vals)

        pre    = best_for(TestResult.PHASE_PRE)
        mid    = best_for(TestResult.PHASE_MID)
        before = best_for(TestResult.PHASE_BEFORE)

        improvement_text = None
        if pre is not None and before is not None:
            unit = (test.unit or '').strip().lower()
            if unit == 's':
                diff = pre - before      # positive = improved (faster)
                improved = diff > 0
            else:
                diff = before - pre      # positive = improved (greater)
                improved = diff > 0
            if diff == 0:
                improvement_text = f"{boxer.name} has no change before the tournament."
            elif improved:
                improvement_text = f"{boxer.name} has improved by {abs(diff)} {test.unit} before the tournament."
            else:
                improvement_text = f"{boxer.name} is now worse by {abs(diff)} {test.unit}."

        ctx = self.get_context_data(
            form=form, boxer=boxer, test=test,
            pre=pre, mid=mid, before=before,
            improvement_text=improvement_text
        )
        return self.render_to_response(ctx)


# ===== DELETE BOXER =====

@login_required
@require_POST
def delete_boxer(request, pk):
    if request.user.is_superuser:
        qs = Boxer.objects.all()
    else:
        qs = Boxer.objects.filter(coach=request.user)

    boxer = get_object_or_404(qs, pk=pk)
    name = boxer.name
    boxer.delete()
    messages.success(request, f"Deleted boxer: {name}")
    return redirect("boxer_list")

class AttendanceListView(LoginRequiredMixin, ListView):
    model = Attendance
    template_name = "attendance/attendance_list.html"
    context_object_name = "attendances"
    paginate_by = 50

    def base_scope(self):
        user = self.request.user
        qs = (
            Attendance.objects
            .select_related("boxer", "class_template", "class_template__gym")
            .order_by("-date", "boxer__first_name", "boxer__last_name")  # ✅ changed
        )

        if ParentProfile.objects.filter(user=user).exists():
            qs = qs.filter(boxer__parent_profiles__user=user)
        elif not user.is_superuser:
            gym = user_gym(self.request)
            if gym:
                qs = qs.filter(
                    Q(boxer__gym=gym) |
                    Q(boxer__shared_with_gyms=gym) |
                    Q(boxer__coaches=user)
                )
            else:
                qs = qs.filter(boxer__coaches=user)
        return qs

    def get_queryset(self):
        qs = self.base_scope()
        date_str = (self.request.GET.get("date") or "").strip()
        q = (self.request.GET.get("q") or "").strip()
        show_absences = self.request.GET.get("show_absences") == "1"

        # ✅ name search now checks first OR last name
        name_filter = Q()
        if q:
            name_filter = Q(boxer__first_name__icontains=q) | Q(boxer__last_name__icontains=q)

        if date_str and not q:
            qs = qs.filter(date=date_str, is_present=not show_absences)
        elif not date_str and q:
            qs = qs.filter(name_filter, is_present=not show_absences)
        elif date_str and q:
            qs = qs.none()  # handled by sentence mode
        else:
            qs = qs.filter(is_present=not show_absences)

        return qs.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_str = (self.request.GET.get("date") or "").strip()
        q = (self.request.GET.get("q") or "").strip()
        show_absences = self.request.GET.get("show_absences") == "1"

        ctx["q"] = q
        ctx["date"] = date_str
        ctx["today"] = timezone.localdate()
        ctx["show_absences"] = show_absences

        # ✅ presence sentence uses first/last filters now
        if date_str and q:
            scoped = self.base_scope().filter(date=date_str).filter(
                Q(boxer__first_name__icontains=q) | Q(boxer__last_name__icontains=q)
            )
            was_present = scoped.filter(is_present=True).exists()
            excused = scoped.filter(is_present=False, is_excused=True).exists()

            ctx["presence_mode"] = True
            ctx["presence_name"] = q
            ctx["presence_date"] = date_str
            ctx["presence_status"] = "present" if was_present else ("excused" if excused else "absent")
        else:
            ctx["presence_mode"] = False

        return ctx


def _attendance_user_scope(request: HttpRequest):
    """
    Mirror AttendanceListView.base_scope() so edits/deletes use the same permission rules.
    """
    user = request.user
    qs = (
        Attendance.objects
        .select_related("boxer", "class_template", "class_template__gym")
        .order_by("-date", "boxer__first_name", "boxer__last_name")
    )
    if ParentProfile.objects.filter(user=user).exists():
        # parents can *see* their kids’ records; usually they shouldn't edit/delete.
        # We’ll keep them read-only by returning an empty qs for mutations.
        return qs.none()
    elif not user.is_superuser:
        gym = user_gym(request)
        if gym:
            qs = qs.filter(
                Q(boxer__gym=gym) |
                Q(boxer__shared_with_gyms=gym) |
                Q(boxer__coaches=user)
            )
        else:
            qs = qs.filter(boxer__coaches=user)
    return qs


def _next_url(request: HttpRequest, fallback_name="attendance_list"):
    """
    Where to go back after edit/delete. Prefers ?next=…; then HTTP_REFERER; finally the list.
    """
    nxt = request.POST.get("next") or request.GET.get("next")
    if nxt:
        return nxt
    ref = request.META.get("HTTP_REFERER")
    if ref:
        return ref
    from django.urls import reverse
    return reverse(fallback_name)


@require_POST
def attendance_edit(request: HttpRequest, pk: int):
    qs = _attendance_user_scope(request)
    att = get_object_or_404(qs, pk=pk)

    date_str = (request.POST.get("date") or "").strip()
    status = (request.POST.get("status") or "").strip().lower()
    next_url = _next_url(request)

    # validate inputs
    new_date = parse_date(date_str)
    if not new_date:
        messages.error(request, "Invalid date.")
        return redirect(next_url)

    if status not in {"present", "absent", "excused"}:
        messages.error(request, "Invalid status.")
        return redirect(next_url)

    # map status → fields
    if status == "present":
        att.is_present = True
        att.is_excused = False
    elif status == "absent":
        att.is_present = False
        att.is_excused = False
    else:  # excused
        att.is_present = False
        att.is_excused = True

    att.date = new_date

    try:
        with transaction.atomic():
            att.save()  # respects unique_together (boxer, date, class_template)
        messages.success(request, "Attendance updated.")
    except IntegrityError:
        messages.error(
            request,
            "Another attendance for this boxer/class on that date already exists.",
        )

    return redirect(next_url)


@require_POST
def attendance_delete(request: HttpRequest, pk: int):
    qs = _attendance_user_scope(request)
    att = get_object_or_404(qs, pk=pk)
    next_url = _next_url(request)

    att.delete()
    messages.success(request, "Attendance deleted.")
    return redirect(next_url)


def parse_iso_date_or_today(s: str | None) -> date:
    try:
        if s:
            return date.fromisoformat(s)  # expects 'YYYY-MM-DD'
    except Exception:
        pass
    return timezone.localdate()


class MarkAttendanceView(LoginRequiredMixin, TemplateView):
    template_name = "attendance/mark_attendance.html"

    # ---------- helpers ----------
    def _target_date(self):
        """Return the target date as a date object, fallback to today."""
        raw = self.request.GET.get("date") or self.request.POST.get("date")
        parsed = parse_date(raw) if raw else None
        return parsed or timezone.localdate()

    def _selected_class(self, gym):
        """Return the selected ClassTemplate object or None."""
        cid = self.request.GET.get("class_id") or self.request.POST.get("class_id")
        if not cid:
            return None
        try:
            return ClassTemplate.objects.get(id=int(cid), gym=gym)
        except (ClassTemplate.DoesNotExist, ValueError, TypeError):
            return None

    def _scoped_boxers_qs(self, gym, class_obj=None):
        """Restrict boxers to the gym or all if superuser; optionally filter by class enrollment."""
        qs = Boxer.objects.all() if self.request.user.is_superuser else Boxer.objects.filter(gym=gym)
        return qs.filter(enrollments__template=class_obj) if class_obj else qs

    def _preserve_redirect(self, class_obj, target_date):
        """Redirect back to mark_attendance with current filters preserved."""
        url = reverse("mark_attendance")
        params = []
        if class_obj:
            params.append(f"class_id={class_obj.id}")
        if target_date:
            params.append(f"date={target_date.isoformat()}")
        return url + ("?" + "&".join(params) if params else "")

    def _day_bounds(self, d: date):
        """Return timezone-aware [start, end) datetimes for day d."""
        tz = timezone.get_current_timezone()
        start = datetime.combine(d, dt_time.min)
        end = start + timedelta(days=1)
        if timezone.is_naive(start):
            start = timezone.make_aware(start, tz)
            end = timezone.make_aware(end, tz)
        return start, end

    def _excused_field_name(self):
        """Return the field name used for excused on Attendance, or None."""
        for fname in ("is_excused", "excused", "excused_absence"):
            try:
                Attendance._meta.get_field(fname)
                return fname
            except Exception:
                continue
        return None

    # ---------- GET ----------
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)
        target_date = self._target_date()
        selected_class = self._selected_class(gym)

        # All classes for selector
        classes = ClassTemplate.objects.filter(gym=gym).order_by("title")

        # Boxers scoped to this gym (and optionally by class enrollment)
        boxers = self._scoped_boxers_qs(gym, selected_class).order_by("first_name", "last_name")

        # Attendance for this date+class
        att_qs = Attendance.objects.filter(
            boxer__in=boxers, date=target_date, class_template=selected_class
        )

        # Determine excused field (if any)
        excused_field = self._excused_field_name()

        # Build map boxer_id -> {status, excused}
        attendance_map = {}
        for att in att_qs.select_related("boxer"):
            excused = bool(getattr(att, excused_field)) if excused_field else False
            status = "present" if att.is_present else ("excused" if excused else "absent")
            attendance_map[att.boxer_id] = {"status": status, "excused": excused}

        # Data for template cards
        boxer_data = []
        for boxer in boxers:
            att = attendance_map.get(boxer.id)
            boxer_data.append({
                "boxer": boxer,
                "status": att["status"] if att else None,   # None = blank if no record yet
                "excused": att["excused"] if att else False,
            })

        ctx.update({
            "date": target_date,
            "classes": classes,
            "selected_class": selected_class,
            "boxers": boxers,  # used for enrolled list
            "boxer_data": boxer_data,  # used for attendance cards
            "class_create_form": ClassCreateForm(request=self.request),
            "enroll_form": EnrollBoxerForm(request=self.request, template=selected_class) if selected_class else None,
            "unenroll_form": UnenrollForm(),
        })
        return ctx

    # ---------- POST ----------
    def post(self, request, *args, **kwargs):
        gym = user_gym(request)
        target_date = self._target_date()
        selected_class = self._selected_class(gym)
        action = request.POST.get("action")

        # A) Create class
        if action == "create_class":
            form = ClassCreateForm(request.POST, request=request)
            if form.is_valid():
                new_cls = form.save()
                return redirect(self._preserve_redirect(new_cls, target_date))
            return self.get(request, *args, **kwargs)

        # B) Enroll boxer
        if action == "enroll" and selected_class:
            form = EnrollBoxerForm(request.POST, request=request, template=selected_class)
            if form.is_valid():
                Enrollment.objects.get_or_create(template=selected_class, boxer=form.cleaned_data["boxer"])
            return redirect(self._preserve_redirect(selected_class, target_date))

        # C) Unenroll boxer
        if action == "unenroll" and selected_class:
            try:
                boxer_id = int(request.POST.get("boxer_id"))
            except (TypeError, ValueError):
                return redirect(self._preserve_redirect(selected_class, target_date))
            Enrollment.objects.filter(template=selected_class, boxer_id=boxer_id).delete()
            return redirect(self._preserve_redirect(selected_class, target_date))

        # C.5) Mark all UNMARKED as Absent (server-side, DB is the source of truth)
        if action == "mark_all_unmarked_absent":
            if not selected_class:
                messages.error(request, "Please select a class first.")
                return redirect(self._preserve_redirect(None, target_date))

            excused_field = self._excused_field_name()
            boxer_qs = self._scoped_boxers_qs(gym, selected_class)

            # Boxer IDs with an existing attendance row for this day+class
            existing_ids = set(
                Attendance.objects.filter(
                    boxer__in=boxer_qs, date=target_date, class_template=selected_class
                ).values_list("boxer_id", flat=True)
            )

            # Create Absent rows for those without any record yet
            to_create = []
            for b in boxer_qs.only("id"):
                if b.id in existing_ids:
                    continue
                row_kwargs = {
                    "boxer": b,
                    "date": target_date,
                    "class_template": selected_class,
                    "is_present": False,
                }
                if excused_field:
                    row_kwargs[excused_field] = False
                to_create.append(Attendance(**row_kwargs))

            if to_create:
                Attendance.objects.bulk_create(to_create, ignore_conflicts=True)
                messages.success(request, "All unmarked boxers were set to Absent.")
            else:
                messages.info(request, "No unmarked boxers to set as Absent.")

            return redirect(self._preserve_redirect(selected_class, target_date))

        # D) Save attendance (individual entries)
        measured_dt = datetime.combine(target_date, dt_time(hour=12, minute=0))
        if timezone.is_naive(measured_dt):
            measured_dt = timezone.make_aware(measured_dt, timezone.get_current_timezone())

        excused_field = self._excused_field_name()

        boxer_qs = self._scoped_boxers_qs(gym, selected_class)
        if not selected_class:
            messages.error(request, "Please select a class before saving attendance.")
            return redirect(self._preserve_redirect(None, target_date))

        for boxer in boxer_qs:
            status = request.POST.get(f"attendance_{boxer.id}")  # "Present" | "Absent" | None
            raw_weight = (request.POST.get(f"weight_{boxer.id}") or "").strip()
            has_weight_input = bool(raw_weight)
            excused_flag = f"excused_{boxer.id}" in request.POST

            if status or has_weight_input:
                # --- Save Attendance ---
                is_present = status == "Present" or (status is None and has_weight_input)
                defaults = {"is_present": is_present}
                if excused_field:
                    defaults[excused_field] = bool(excused_flag) and not is_present

                Attendance.objects.update_or_create(
                    boxer=boxer,
                    date=target_date,
                    class_template=selected_class,
                    defaults=defaults,
                )

                # --- Save Weight (only if explicitly entered) ---
                if is_present and has_weight_input:
                    try:
                        kg = Decimal(raw_weight)
                    except (InvalidOperation, ValueError):
                        kg = None
                    if kg is not None:
                        Weight.objects.update_or_create(
                            boxer=boxer,
                            measured_at=measured_dt,
                            defaults={"kg": kg},
                        )
            else:
                # If nothing selected and no weight → remove any existing record
                Attendance.objects.filter(
                    boxer=boxer,
                    date=target_date,
                    class_template=selected_class
                ).delete()

        return redirect(self._preserve_redirect(selected_class, target_date))


@login_required
def attendance_by_date(request):
    # parse ?date=YYYY-MM-DD; default to today if missing/invalid
    q = (request.GET.get("date") or "").strip()
    try:
        selected = Date.fromisoformat(q) if q else timezone.now().date()
    except ValueError:
        selected = timezone.now().date()

    # scope boxers based on current user
    gym = user_gym(request)
    if request.user.is_superuser:
        boxers_qs = Boxer.objects.all()
    elif gym:
        boxers_qs = Boxer.objects.filter(
            Q(gym=gym) | Q(shared_with_gyms=gym) | Q(coaches=request.user)
        ).distinct()
        if not boxers_qs.exists():
            boxers_qs = Boxer.objects.filter(coaches=request.user).distinct() or Boxer.objects.all()
    else:
        boxers_qs = Boxer.objects.filter(coaches=request.user).distinct() or Boxer.objects.all()

    records = (
        Attendance.objects
        .filter(boxer__in=boxers_qs, date=selected)
        .select_related("boxer")
        .order_by("boxer__name")
    )

    return render(request, "attendance/attendance_by_date.html", {
        "date": selected,                  # for display
        "date_value": selected.isoformat(),  # for input value
        "records": records,
    })
@login_required
@require_POST
def delete_attendance(request, attendance_id):
    record = get_object_or_404(Attendance, id=attendance_id, boxer__coach=request.user)
    record.delete()
    messages.success(request, "Attendance record deleted.")
    return redirect('attendance_list')

class BoxerResumeView(LoginRequiredMixin, TemplateView):
    template_name = "boxers/boxer_resume.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        boxer = get_object_or_404(Boxer, pk=self.kwargs["boxer_id"])
        ctx["boxer"] = boxer

        # Attendance
        att_qs = Attendance.objects.filter(boxer=boxer)
        att_total = att_qs.count()
        att_present = att_qs.filter(is_present=True).count()
        att_absent = att_total - att_present
        att_excused = att_qs.filter(is_present=False, is_excused=True).count()

        def pct(n, d):
            return round((n * 100.0) / d, 1) if d else 0.0

        ctx["att_total"] = att_total
        ctx["att_present"] = att_present
        ctx["att_absent"] = att_absent
        ctx["att_excused"] = att_excused
        ctx["att_present_pct"] = pct(att_present, att_total)
        ctx["att_absent_pct"] = pct(att_absent, att_total)
        ctx["att_excused_pct"] = pct(att_excused, att_absent)

        # Weight
        weights = Weight.objects.filter(boxer=boxer).order_by("measured_at")
        if weights.exists():
            ctx["weight_min"] = weights.order_by("kg").first()
            ctx["weight_max"] = weights.order_by("-kg").first()
            ctx["weight_avg"] = round(sum(w.kg for w in weights) / weights.count(), 2)
        else:
            ctx["weight_min"] = None
            ctx["weight_max"] = None
            ctx["weight_avg"] = None

        return ctx



class BoxerReportView(LoginRequiredMixin, TemplateView):
    template_name = "boxers/boxer_report.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        boxer_id = self.kwargs.get("boxer_id")
        boxer = get_object_or_404(Boxer, pk=boxer_id)
        att_qs = Attendance.objects.filter(boxer=boxer).order_by("-date")

        # Attendance stats
        excused_field = None
        for fname in ("is_excused", "excused", "excused_absence"):
            if hasattr(Attendance, fname):
                excused_field = fname
                break

        att_total = att_qs.count()
        att_present = att_qs.filter(is_present=True).count()
        att_absent = att_qs.filter(is_present=False).count()
        att_excused = att_qs.filter(**{excused_field: True}).count() if excused_field else 0

        def pct(n, d): return round((n * 100.0) / d, 1) if d else 0.0

        att_present_pct = pct(att_present, att_total)
        att_absent_pct = pct(att_absent, att_total)
        att_excused_pct_of_abs = pct(att_excused, att_absent)

        # Daily weights
        weight_by_day = {}
        for w in Weight.objects.filter(boxer=boxer).order_by("measured_at"):
            try:
                day = w.measured_at.date()
            except Exception:
                continue
            weight_by_day[day] = w.kg

        rows = []
        for a in att_qs:
            rows.append({
                "date": a.date,
                "is_present": a.is_present,
                "is_excused": (getattr(a, excused_field, False) if excused_field else False),
                "weight_kg": weight_by_day.get(a.date),
            })

        # Weight stats
        weight_qs = Weight.objects.filter(boxer=boxer)
        if weight_qs.exists():
            weight_stats = {
                "min": weight_qs.aggregate(Min("kg"))["kg__min"],
                "min_date": weight_qs.order_by("kg").first().measured_at.date(),
                "max": weight_qs.aggregate(Max("kg"))["kg__max"],
                "max_date": weight_qs.order_by("-kg").first().measured_at.date(),
                "avg": round(weight_qs.aggregate(Avg("kg"))["kg__avg"], 1),
            }
        else:
            weight_stats = None

        ctx.update({
            "boxer": boxer,
            "attendance": rows,
            "att_total": att_total,
            "att_present": att_present,
            "att_absent": att_absent,
            "att_excused": att_excused,
            "att_present_pct": att_present_pct,
            "att_absent_pct": att_absent_pct,
            "att_excused_pct_of_abs": att_excused_pct_of_abs,
            "weight_stats": weight_stats,
            "today": timezone.localdate(),
        })
        return ctx

class BoxerClassesView(LoginRequiredMixin, DetailView):
    template_name = "boxers/boxer_classes.html"
    model = Boxer
    context_object_name = "boxer"
    pk_url_kwarg = "boxer_id"

    def get_queryset(self):
        # keep gym scoping similar to the rest of your app
        if self.request.user.is_superuser:
            return Boxer.objects.all()
        gym = user_gym(self.request)
        return Boxer.objects.filter(gym=gym)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        boxer = self.object
        ctx["classes"] = (
            ClassTemplate.objects
            .filter(enrollments__boxer=boxer)
            .order_by("title")
        )
        return ctx

@login_required
def add_boxer(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Please enter a boxer name.")
            return redirect("home")

        # Create boxer without coach
        boxer = Boxer.objects.create(name=name, gym=user_gym(request))

        # Attach the current user as one of the coaches
        boxer.coaches.add(request.user)

        messages.success(
            request,
            f"You successfully added {boxer.name} to your boxers' list."
        )
        return redirect("home")

    return redirect("home")

@login_required
@require_POST
def summary_result_save(request):
    # Parse required ids + phase
    try:
        boxer_id = int(request.POST.get('boxer_id'))
        test_id  = int(request.POST.get('test_id'))
        phase    = request.POST.get('phase') or TestResult.PHASE_PRE
    except (TypeError, ValueError):
        return HttpResponseBadRequest("Invalid parameters")

    # Ownership checks
    boxer = get_object_or_404(Boxer, id=boxer_id, coach=request.user)
    test  = get_object_or_404(BatteryTest, id=test_id, coach=request.user)

    # Parse optional numeric fields
    def parse_val(name):
        raw = (request.POST.get(name) or '').strip()
        if raw == '':
            return None
        try:
            return float(raw)
        except ValueError:
            raise ValidationError(f"{name} must be a number or blank.")

    notes = (request.POST.get('notes') or '').strip()
    try:
        v1 = parse_val('value1')
        v2 = parse_val('value2')
        v3 = parse_val('value3')
    except ValidationError as e:
        messages.error(request, str(e))
        return redirect(request.META.get('HTTP_REFERER', reverse('tests_summary')) + f'?phase={phase}')

    # Create/update and save
    obj, created = TestResult.objects.get_or_create(boxer=boxer, test=test, phase=phase)
    obj.value1, obj.value2, obj.value3, obj.notes = v1, v2, v3, notes
    obj.save()

    messages.success(request, "Result saved.")
    return redirect(reverse('tests_summary') + f'?phase={phase}')

@login_required
@require_POST
def summary_result_delete(request):
    try:
        boxer_id = int(request.POST.get('boxer_id'))
        test_id  = int(request.POST.get('test_id'))
        phase    = (request.POST.get('phase') or TestResult.PHASE_PRE)
    except Exception:
        return HttpResponseBadRequest("Invalid parameters")

    boxer = get_object_or_404(Boxer, id=boxer_id, coach=request.user)
    test  = get_object_or_404(BatteryTest, id=test_id, coach=request.user)

    TestResult.objects.filter(boxer=boxer, test=test, phase=phase).delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})

    messages.success(request, "Result deleted.")
    return redirect(reverse('tests_summary') + f'?phase={phase}')


class TestResultEditView(LoginRequiredMixin, TemplateView):
    """Tiny page to edit a single cell (useful if you don't want inline inputs)."""
    template_name = 'tests/tests_result_edit.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        request = self.request
        boxer_id = int(request.GET.get('boxer'))
        test_id  = int(request.GET.get('test'))
        phase    = request.GET.get('phase') or TestResult.PHASE_PRE
        boxer = get_object_or_404(Boxer, id=boxer_id, coach=request.user)
        test  = get_object_or_404(BatteryTest, id=test_id, coach=request.user)
        result = TestResult.objects.filter(boxer=boxer, test=test, phase=phase).first()
        ctx.update({'boxer': boxer, 'test': test, 'phase': phase, 'result': result})
        return ctx

    def post(self, request, *args, **kwargs):
        # Reuse the save logic (handles both JSON and regular form posts)
        return summary_result_save(request)

@login_required
def record_heart_rate(request):
    if request.method != "POST":
        return redirect("home")

    form = HeartRateQuickForm(request.POST, user=request.user)
    if not form.is_valid():
        messages.error(request, "Please correct the heart rate form.")
        return redirect("home")

    boxer = form.cleaned_data["boxer"]
    # safety: ensure the boxer belongs to the logged-in coach
    boxer = get_object_or_404(Boxer, id=boxer.id, coach=request.user)

    phase = form.cleaned_data["phase"]
    bpm = form.cleaned_data["bpm"]
    measured_at = form.cleaned_data["measured_at"] or timezone.now().date()

    obj, created = HeartRate.objects.get_or_create(
        boxer=boxer, phase=phase,
        defaults={"bpm": bpm, "measured_at": measured_at}
    )
    if not created:
        obj.bpm = bpm
        obj.measured_at = measured_at
        obj.save()

    messages.success(request, f"Saved {boxer.name} – {dict(TestResult.PHASE_CHOICES).get(phase, phase)}: {bpm} bpm.")
    return redirect("home")


@login_required
def record_heart_rate(request):
    if request.method != "POST":
        return redirect("home")

    form = HeartRateQuickForm(request.POST)
    if form.is_valid():
        hr = form.save(commit=False)
        # (Optional) gate: ensure boxer is in scope of current user
        gym = user_gym(request)
        allowed = (
            request.user.is_superuser
            or (gym and hr.boxer.gym_id == getattr(gym, "id", None))
            or (gym and hr.boxer.shared_with_gyms.filter(id=gym.id).exists())
            or hr.boxer.coaches.filter(id=request.user.id).exists()
        )
        if not allowed:
            return redirect("home")
        hr.save()
    return redirect("heart_rate_summary")


class HeartRateSummaryView(LoginRequiredMixin, TemplateView):
    template_name = "heartrate/heart_rate_summary.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)

        # Scope boxers like HomeView
        if self.request.user.is_superuser:
            boxers = Boxer.objects.all()
        elif gym:
            boxers = Boxer.objects.filter(
                Q(gym=gym) | Q(shared_with_gyms=gym) | Q(coaches=self.request.user)
            ).distinct()
        else:
            boxers = Boxer.objects.filter(coaches=self.request.user).distinct()
        boxers = boxers.order_by("name")

        # Pull heart rates for those boxers
        hr_qs = HeartRate.objects.filter(boxer__in=boxers)

        # Find the latest timestamp per boxer
        latest = hr_qs.values("boxer_id").annotate(last_measured=Max("measured_at"))
        last_map = {row["boxer_id"]: row["last_measured"] for row in latest}

        # Fetch actual HeartRate rows at those timestamps
        filt = Q()
        for b_id, ts in last_map.items():
            filt |= (Q(boxer_id=b_id) & Q(measured_at=ts))
        latest_objs = hr_qs.filter(filt) if last_map else hr_qs.none()
        latest_objs = latest_objs.select_related("boxer")

        # Map boxer_id -> HeartRate object
        obj_map = {hr.boxer_id: hr for hr in latest_objs}

        # Build rows for template
        rows = []
        for boxer in boxers:
            rows.append({
                "boxer": boxer,
                "latest": obj_map.get(boxer.id),
            })

        ctx["rows"] = rows
        return ctx

def _scoped_boxers(request):
    user = request.user
    if user.is_superuser:
        return Boxer.objects.all()
    cp = getattr(user, "coachprofile", None)
    gym = getattr(cp, "gym", None)
    if gym:
        return Boxer.objects.filter(Q(coaches=user) | Q(gym=gym) | Q(shared_with_gyms=gym)).distinct()
    return Boxer.objects.filter(coaches=user).distinct()

class HeartRateDetailView(LoginRequiredMixin, DetailView):
    model = Boxer
    template_name = "heartrate/boxer_detail.html"
    pk_url_kwarg = "boxer_id"
    def get_queryset(self):
        return _scoped_boxers(self.request)

class HeartRateCreateView(LoginRequiredMixin, CreateView):
    form_class = HeartRateQuickForm
    template_name = "heartrate/heart_rate_form.html"

    def dispatch(self, request, *args, **kwargs):
        self.boxer = get_object_or_404(_scoped_boxers(request), pk=kwargs["boxer_id"])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["boxer"] = self.boxer
        return ctx

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.boxer = self.boxer
        obj.save()
        messages.success(self.request, f"Added {obj.bpm} bpm for {self.boxer.name}.")
        next_url = self.request.GET.get("next")
        return redirect(next_url) if next_url else redirect(
            reverse("heart_rate_detail", kwargs={"boxer_id": self.boxer.id})
        )

@login_required
def record_weight(request):
    if request.method != "POST":
        return redirect("home")

    form = WeightQuickForm(request.POST, user=request.user)
    if not form.is_valid():
        messages.error(request, "Please correct the weight form.")
        return redirect("home")

    boxer = get_object_or_404(Boxer, id=form.cleaned_data["boxer"].id, coach=request.user)
    phase = form.cleaned_data["phase"]
    kg = form.cleaned_data["kg"]
    expected_kg = form.cleaned_data.get("expected_kg")
    measured_at = form.cleaned_data.get("measured_at") or timezone.now().date()

    obj, created = Weight.objects.get_or_create(
        boxer=boxer, phase=phase,
        defaults={"kg": kg, "expected_kg": expected_kg, "measured_at": measured_at}
    )
    if not created:
        obj.kg = kg
        obj.expected_kg = expected_kg
        obj.measured_at = measured_at
        obj.save()

    phase_label = dict(TestResult.PHASE_CHOICES).get(phase, phase)
    messages.success(request, f"Saved {boxer.name} — {phase_label}: {kg} kg (expected: {expected_kg or 'n/a'}).")
    return redirect("home")

class WeightSummaryView(LoginRequiredMixin, TemplateView):
    template_name = 'weight/weight_summary.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        boxers = list(Boxer.objects.filter(coach=self.request.user).order_by('name'))
        weights = Weight.objects.filter(boxer__coach=self.request.user)
        wmap = {(w.boxer_id, w.phase): w for w in weights}
        rows = []
        for b in boxers:
            rows.append({
                "boxer": b,
                "pre":    wmap.get((b.id, TestResult.PHASE_PRE)),
                "mid":    wmap.get((b.id, TestResult.PHASE_MID)),
                "before": wmap.get((b.id, TestResult.PHASE_BEFORE)),
            })
        ctx["rows"] = rows
        return ctx

class WeightDetailView(LoginRequiredMixin, TemplateView):
    template_name = "weight/weight_detail.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        boxer = get_object_or_404(Boxer, id=self.kwargs["boxer_id"], coach=self.request.user)
        weights = {w.phase: w for w in Weight.objects.filter(boxer=boxer)}
        ctx.update({
            "boxer": boxer,
            "pre":    weights.get(TestResult.PHASE_PRE),
            "mid":    weights.get(TestResult.PHASE_MID),
            "before": weights.get(TestResult.PHASE_BEFORE),
        })
        return ctx


class TestRankingView(LoginRequiredMixin, TemplateView):
    template_name = "tests/tests_rankings.html"

    def _scoped_boxers(self, user, gym):
        if user.is_superuser:
            return Boxer.objects.all()
        qs = Boxer.objects.filter(coaches=user).distinct()
        if gym:
            qs = Boxer.objects.filter(
                Q(coaches=user) | Q(gym=gym) | Q(shared_with_gyms=gym)
            ).distinct()
        return qs if qs.exists() else Boxer.objects.all()

    def _lower_is_better(self, test):
        """Return True if lower values should rank higher for this test."""
        u = (getattr(test, "unit", "") or "").strip().lower()
        time_units = {
            "s", "sec", "secs", "second", "seconds",
            "ms", "millisecond", "milliseconds",
            "min", "mins", "minute", "minutes",
            "h", "hr", "hrs", "hour", "hours",
        }
        if u in time_units:
            return True
        if u in {"m", "meter", "meters", "metre", "metres", "cm", "centimeter",
                 "centimeters", "centimetre", "centimetres"}:
            return False
        return False  # default: higher is better

    def get(self, request, *args, **kwargs):
        gym = user_gym(request)
        boxers_qs = self._scoped_boxers(request.user, gym).order_by("name")

        tests = BatteryTest.objects.all().order_by("display_order", "name")

        # Pick selected test (by URL kwarg or ?test=) or first test that has results
        selected_test = None
        test_id = kwargs.get("test_id") or request.GET.get("test")
        if test_id:
            selected_test = get_object_or_404(BatteryTest, pk=int(test_id))
        else:
            exists_qs = TestResult.objects.filter(test=OuterRef("pk"), boxer__in=boxers_qs)
            selected_test = (
                tests.annotate(has_results=Exists(exists_qs))
                     .filter(has_results=True)
                     .first()
            )

        rows = []
        if selected_test:
            res_qs = TestResult.objects.filter(test=selected_test, boxer__in=boxers_qs)

            lower_is_better = self._lower_is_better(selected_test)

            by_boxer = {}
            for r in res_qs.select_related("boxer"):
                vals = []
                for v in (r.value1, r.value2, r.value3):
                    try:
                        if v is not None:
                            vals.append(Decimal(str(v)))
                    except Exception:
                        pass
                if not vals:
                    continue

                best = min(vals) if lower_is_better else max(vals)
                cur = by_boxer.get(r.boxer_id)
                if cur is None or (lower_is_better and best < cur["best"]) or (not lower_is_better and best > cur["best"]):
                    by_boxer[r.boxer_id] = {"boxer": r.boxer, "best": best}

            rows = sorted(by_boxer.values(), key=lambda d: d["best"], reverse=not lower_is_better)

        ctx = {
            "tests": tests,
            "selected_test": selected_test,
            "rows": rows,
        }
        return self.render_to_response(ctx)



def health(request):
    return HttpResponse("ok")

def debug_urls(request):
    pats = [str(p.pattern) for p in get_resolver().url_patterns]
    return JsonResponse({"patterns": pats})

def debug_env(request):
    from django.conf import settings
    return JsonResponse({
        "ALLOWED_HOSTS": settings.ALLOWED_HOSTS,
        "CSRF_TRUSTED_ORIGINS": getattr(settings, "CSRF_TRUSTED_ORIGINS", []),
        "DEBUG": settings.DEBUG,
    })


@staff_member_required
def export_fixture(request):
    buf = io.StringIO()
    call_command(
        "dumpdata",
        "--natural-foreign", "--natural-primary",
        "-e", "contenttypes", "-e", "auth.permission", "-e", "admin.logentry", "-e", "sessions",
        stdout=buf,
    )
    data = buf.getvalue()
    resp = HttpResponse(data, content_type="application/json")
    resp["Content-Disposition"] = 'attachment; filename="render_dump.json"'
    return resp


def user_can_view_boxer(user, boxer):
    """Allow superuser, the coach who owns the boxer, or a parent linked to the boxer."""
    if not user.is_authenticated:
        return False
    if user.is_superuser or getattr(boxer, 'coach_id', None) == user.id:
        return True
    try:
        return user.parent_profile.children.filter(id=boxer.id).exists()
    except ParentProfile.DoesNotExist:
        return False


class ParentHomeView(LoginRequiredMixin, TemplateView):
    template_name = "parent/home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Example scope: all boxers in the user's gym (adjust to your needs)
        gym = user_gym(self.request)

        qs = Boxer.objects.all()
        if not self.request.user.is_superuser and gym:
            qs = qs.filter(gym=gym)

        # ✅ Valid relational loading:
        qs = qs.select_related("gym").prefetch_related("coaches")  # or "gym__coaches" if coaches are on Gym

        ctx["boxers"] = qs.order_by("name")
        return ctx

class ParentAttendanceView(LoginRequiredMixin, TemplateView):
    template_name = 'parent/attendance.html'

    def get(self, request, boxer_id, *args, **kwargs):
        boxer = get_object_or_404(Boxer, id=boxer_id)
        if not user_can_view_boxer(request.user, boxer):
            raise PermissionDenied("You cannot view this boxer.")
        records = Attendance.objects.filter(boxer=boxer).order_by('-date')
        return render(request, self.template_name, {'boxer': boxer, 'records': records})


class WeightProgressView(LoginRequiredMixin, TemplateView):
    template_name = "weight/weight_progress.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["attendance_fallback"] = reverse("attendance_list")
        boxer = get_object_or_404(Boxer, pk=self.kwargs["boxer_id"])

        gym = user_gym(self.request)
        allowed = (
            self.request.user.is_superuser
            or (gym and boxer.gym_id == getattr(gym, "id", None))
            or (gym and boxer.shared_with_gyms.filter(id=gym.id).exists())
            or boxer.coaches.filter(id=self.request.user.id).exists()
            or ParentProfile.objects.filter(user=self.request.user, children=boxer).exists()
        )
        if not allowed:
            ctx.update({"boxer": boxer, "rows": [], "fw_error": "Not allowed"})
            return ctx

        # Take the latest entry per calendar date
        by_day = {}
        for w in Weight.objects.filter(boxer=boxer).order_by("measured_at"):
            day = w.measured_at.date()
            by_day[day] = w.kg

        rows = [{"date": d, "weight": by_day[d]} for d in sorted(by_day)]

        vals = [r["weight"] for r in rows if r["weight"] is not None]
        min_w = min(vals) if vals else None
        max_w = max(vals) if vals else None
        diff = (max_w - min_w) if (min_w is not None and max_w is not None) else None
        avg_w = (sum(vals) / len(vals)) if vals else None  # ✅ average

        fw_raw = (self.request.GET.get("fighting_weight") or "").strip()
        fighting_weight, fw_error, count_above_fw = None, None, None
        if fw_raw:
            try:
                fighting_weight = Decimal(fw_raw)
                count_above_fw = sum(1 for v in vals if v is not None and v > fighting_weight)
            except (InvalidOperation, TypeError):
                fw_error = "Invalid fighting weight. Please enter a number (e.g., 72.5)."

        ctx.update({
            "boxer": boxer,
            "rows": rows,
            "min_w": min_w,
            "max_w": max_w,
            "diff": diff,
            "avg_w": avg_w,                     # ✅ send average
            "fighting_weight": fighting_weight,
            "count_above_fw": count_above_fw,
            "fw_error": fw_error,
        })
        return ctx

class ParentSignupView(FormView):
    template_name = 'parent/signup.html'
    form_class = ParentSignupForm
    success_url = reverse_lazy('parent_home')

    def form_valid(self, form):
        # Create the user with hashed password (UserCreationForm handles hashing)
        user = form.save()
        # Ensure it is a regular, non-staff account
        user.is_staff = False
        user.is_superuser = False
        user.save(update_fields=["is_staff", "is_superuser"])

        # Link to the selected child
        child = form.cleaned_data["child"]
        pp, _ = ParentProfile.objects.get_or_create(user=user)
        pp.children.add(child)

        # Log them in and send to parent portal
        login(self.request, user)
        return super().form_valid(form)


class CalendarView(TemplateView):
    template_name = "calendar.html"

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return HttpResponseForbidden()
        # You can preload templates list if you want to show a sidebar
        templates = ClassTemplate.objects.filter(gym=user_gym(request)).order_by("title")
        return render(request, self.template_name, {"templates": templates})

@login_required
def api_class_attendance(request):
    gym = user_gym(request)
    start = parse_date(request.GET.get("start"))
    end   = parse_date(request.GET.get("end"))
    if not start or not end:
        return HttpResponseBadRequest("start and end (YYYY-MM-DD) required")

    templates = ClassTemplate.objects.filter(gym=gym)

    payload = []
    for t in templates:
        # get distinct dates where attendance exists in range
        dates = (
            Attendance.objects
            .filter(class_template=t, date__gte=start, date__lte=end)
            .values_list("date", flat=True)
            .distinct()
        )

        for d in dates:
            present   = Attendance.objects.filter(class_template=t, date=d, is_present=True).count()
            enrolled  = Enrollment.objects.filter(template=t).count()

            payload.append({
                "class_id": t.id,
                "class_title": t.title,
                "date": d.isoformat(),
                "present_count": present,
                "enrolled_count": enrolled,
            })

    return JsonResponse({"classes": payload})



@login_required
@transaction.atomic
def api_enroll(request):
    gym = user_gym(request)
    if request.method == "POST":
        # Add enrollment (roster add)
        template_id = request.POST.get("template_id")
        boxer_id = request.POST.get("boxer_id")
        effective_from = parse_date(request.POST.get("effective_from")) or date.today()

        template = get_object_or_404(ClassTemplate, id=template_id, gym=gym)
        boxer = get_object_or_404(Boxer, id=boxer_id, gym=gym)

        Enrollment.objects.create(template=template, boxer=boxer, active_from=effective_from)
        return JsonResponse({"ok": True})

    if request.method == "DELETE":
        # End enrollment (roster remove)
        template_id = request.GET.get("template_id")
        boxer_id = request.GET.get("boxer_id")
        effective_until = parse_date(request.GET.get("effective_until")) or date.today()

        e = get_object_or_404(Enrollment, template_id=template_id, boxer_id=boxer_id)
        e.active_until = effective_until
        e.save(update_fields=["active_until"])
        return JsonResponse({"ok": True})

    return HttpResponseBadRequest("POST to add, DELETE to remove")


@login_required
@transaction.atomic
def api_attendance_upsert(request):
    gym = user_gym(request)
    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    class_id = request.POST.get("class_id")
    boxer_id = request.POST.get("boxer_id")
    date_raw = request.POST.get("date") or timezone.localdate().isoformat()
    status   = request.POST.get("status")    # 'present' | 'absent' | 'excused'
    weight_raw = request.POST.get("weight")

    class_template = get_object_or_404(ClassTemplate, id=class_id, gym=gym)
    boxer = get_object_or_404(Boxer, id=boxer_id, gym=gym)

    try:
        date_val = date.fromisoformat(date_raw)
    except Exception:
        return HttpResponseBadRequest("Invalid date")

    # Upsert attendance
    att, _ = Attendance.objects.get_or_create(
        class_template=class_template,
        boxer=boxer,
        date=date_val,
        defaults={"is_present": False, "is_excused": False},
    )

    if status == "present":
        att.is_present = True
        att.is_excused = False
    elif status == "excused":
        att.is_present = False
        att.is_excused = True
    elif status == "absent":
        att.is_present = False
        att.is_excused = False

    if weight_raw not in (None, ""):
        try:
            weight_val = Decimal(weight_raw)
        except Exception:
            return HttpResponseBadRequest("Invalid weight")
        att.is_present = True
        att.is_excused = False
        # Optionally: store weight on Weight model instead of Attendance
        Weight.objects.update_or_create(
            boxer=boxer,
            measured_at=datetime.combine(date_val, dt_time(hour=12, minute=0)),
            defaults={"kg": weight_val},
        )

    att.save()

    return JsonResponse({
        "ok": True,
        "is_present": att.is_present,
        "is_excused": att.is_excused,
    })


class GymCreateView(LoginRequiredMixin, CreateView):
    model = Gym
    form_class = GymForm
    template_name = "gyms/gym_form.html"
    success_url = reverse_lazy("gym_list")

class GymListView(LoginRequiredMixin, ListView):
    model = Gym
    template_name = "gyms/gym_list.html"
    context_object_name = "gyms"



def _back_to_tests_list():
    return redirect("tests_list")




def lower_is_better(test):
    """Return True if lower values should rank higher for this test."""
    u = (getattr(test, "unit", "") or "").strip().lower()
    time_units = {
        "s", "sec", "secs", "second", "seconds",
        "ms", "millisecond", "milliseconds",
        "min", "mins", "minute", "minutes",
        "h", "hr", "hrs", "hour", "hours",
    }
    if u in time_units:
        return True
    if u in {"m", "meter", "meters", "metre", "metres",
             "cm", "centimeter", "centimeters",
             "centimetre", "centimetres"}:
        return False
    return False


class TestResultCreateView(LoginRequiredMixin, CreateView):
    template_name = "tests/tests_results.html"   # NEW template name
    form_class = TestResultForm
    success_url = reverse_lazy("tests_record")

    def _selected_class(self, gym):
        raw = (self.request.GET.get("class_id") or "").strip()
        try:
            cid = int(raw)
        except (TypeError, ValueError):
            return None
        try:
            return ClassTemplate.objects.get(id=cid, gym=gym)
        except ClassTemplate.DoesNotExist:
            return None

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        gym = user_gym(self.request)

        # Limit tests (you can scope by gym if needed)
        form.fields["test"].queryset = form.fields["test"].queryset.order_by("display_order", "name")

        # Filter boxers based on class_id (GET)
        selected_class = self._selected_class(gym)
        if selected_class:
            boxers_qs = Boxer.objects.filter(
                gym=gym,
                enrollments__template=selected_class
            ).distinct()
        else:
            boxers_qs = Boxer.objects.filter(gym=gym)

        # Nice ordering; fallback to 'name' for legacy data
        form.fields["boxer"].queryset = boxers_qs.order_by("first_name", "last_name", "name")
        form.fields["boxer"].required = True
        form.fields["test"].required = True

        # Optional preselects
        if bid := self.request.GET.get("boxer"):
            form.initial["boxer"] = bid
        if tid := self.request.GET.get("test"):
            form.initial["test"] = tid

        return form

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)
        classes = ClassTemplate.objects.filter(gym=gym).order_by("title")
        selected_class = self._selected_class(gym)

        # Count visible boxers to show a friendly hint
        boxer_qs = self.get_form().fields["boxer"].queryset
        ctx.update({
            "classes": classes,
            "selected_class": selected_class,
            "visible_boxer_count": boxer_qs.count(),
        })
        return ctx

    def form_valid(self, form):
        resp = super().form_valid(form)
        tr = self.object

        # Collect all values
        vals = [v for v in (tr.value1, tr.value2, tr.value3) if v is not None]

        # Decide best according to test unit
        best = None
        if vals:
            if lower_is_better(tr.test):
                best = min(vals)
            else:
                best = max(vals)

        unit = getattr(tr.test, "unit", "") or ""
        date_str = tr.measured_at.strftime("%Y-%m-%d %H:%M") if tr.measured_at else "unspecified date"

        if best is not None:
            messages.success(self.request, f"{tr.boxer} scored {best} {unit} in '{tr.test}' on {date_str}.")
        else:
            messages.warning(self.request, f"{tr.boxer} had no valid measurement values for '{tr.test}' on {date_str}.")

        return resp

class TestResultBulkCreateView(LoginRequiredMixin, TemplateView):
    template_name = "tests/tests_record_multi.html"

    # --- helpers ---
    def _selected_class(self, gym):
        raw = (self.request.GET.get("class_id") or "").strip()
        try:
            cid = int(raw)
        except (TypeError, ValueError):
            return None
        try:
            return ClassTemplate.objects.get(id=cid, gym=gym)
        except ClassTemplate.DoesNotExist:
            return None

    def _parse_dt(self, value):
        """Parse HTML datetime-local (YYYY-MM-DDTHH:MM) into an aware datetime."""
        if not value:
            return timezone.now()
        try:
            # Handles 'YYYY-MM-DDTHH:MM' and also with :SS if present
            naive = datetime.fromisoformat(value)
        except ValueError:
            # Fallback in case the browser sends a space instead of 'T'
            try:
                naive = datetime.strptime(value, "%Y-%m-%d %H:%M")
            except ValueError:
                return timezone.now()
        # Make timezone-aware in the current TZ
        return timezone.make_aware(naive, timezone.get_current_timezone()) if timezone.is_naive(naive) else naive

    def _to_float(self, s):
        if s in (None, ""):
            return None
        try:
            return float(s)
        except (TypeError, ValueError):
            return None

    # --- GET ---
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        gym = user_gym(self.request)
        selected_class = self._selected_class(gym)

        if selected_class:
            boxers = (Boxer.objects
                      .filter(gym=gym, enrollments__template=selected_class)
                      .distinct()
                      .order_by("first_name", "last_name", "name"))
        else:
            boxers = Boxer.objects.filter(gym=gym).order_by("first_name", "last_name", "name")

        ctx.update({
            "classes": ClassTemplate.objects.filter(gym=gym).order_by("title"),
            "selected_class": selected_class,
            "tests": BatteryTest.objects.order_by("display_order", "name"),
            "boxers": boxers,
            "visible_boxer_count": boxers.count(),
        })
        return ctx

    # --- POST ---
    def post(self, request, *args, **kwargs):
        gym = user_gym(request)
        test_id = request.POST.get("test")
        measured_at_raw = request.POST.get("measured_at")
        phase = request.POST.get("phase") or TestResult.PHASE_PRE
        class_id = request.GET.get("class_id")  # keep the filter after save

        if not test_id or not measured_at_raw:
            messages.error(request, "Please select a test and date.")
            return redirect("tests_record_multi")

        test = get_object_or_404(BatteryTest, id=test_id)
        measured_at = self._parse_dt(measured_at_raw)

        # filter the same set of boxers as in GET
        selected_class = self._selected_class(gym)
        if selected_class:
            boxers_qs = (Boxer.objects
                         .filter(gym=gym, enrollments__template=selected_class)
                         .distinct())
        else:
            boxers_qs = Boxer.objects.filter(gym=gym)

        created = 0
        for boxer in boxers_qs:
            prefix = f"boxer_{boxer.id}_"
            v1 = self._to_float(request.POST.get(prefix + "v1"))
            v2 = self._to_float(request.POST.get(prefix + "v2"))
            v3 = self._to_float(request.POST.get(prefix + "v3"))
            notes = request.POST.get(prefix + "notes") or ""

            # only save if something provided
            if any(val is not None for val in (v1, v2, v3)) or notes:
                TestResult.objects.create(
                    boxer=boxer,
                    test=test,
                    phase=phase,
                    measured_at=measured_at,
                    coach=request.user,
                    value1=v1,
                    value2=v2,
                    value3=v3,
                    notes=notes,
                )
                created += 1

        messages.success(request, f"✅ Recorded {created} results for '{test.name}'.")
        redirect_url = reverse("tests_record_multi")
        if class_id:
            redirect_url += f"?class_id={class_id}"
        return redirect(redirect_url)


class BoxerTestsView(LoginRequiredMixin, TemplateView):
    template_name = "tests/boxer_tests.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        boxer = get_object_or_404(Boxer, uuid=kwargs["uuid"])

        # All available tests + annotate if boxer has results
        tests = (
            BatteryTest.objects
            .annotate(
                has_results=Exists(
                    TestResult.objects.filter(boxer=boxer, test=OuterRef("pk"))
                )
            )
            .order_by("display_order", "name")
        )

        # Which test is selected
        sel_test_id = self.request.GET.get("test")
        selected_test = None
        if sel_test_id:
            selected_test = get_object_or_404(BatteryTest, pk=sel_test_id)
            if not TestResult.objects.filter(boxer=boxer, test=selected_test).exists():
                selected_test = None  # no results, skip chart

        labels, values, summary = [], [], []

        if selected_test:
            qs = (
                TestResult.objects.filter(boxer=boxer, test=selected_test)
                .annotate(day=TruncDate("measured_at"))
                .values("day", "value1", "value2", "value3", "notes")
                .order_by("day")
            )

            grouped = defaultdict(list)
            for row in qs:
                grouped[row["day"]].append(row)

            for d, items in grouped.items():
                nums, notes = [], []
                for item in items:
                    for v in (item["value1"], item["value2"], item["value3"]):
                        if v is not None:
                            nums.append(v)
                    if item["notes"]:
                        notes.append(item["notes"])

                if nums:
                    best_val = min(nums) if lower_is_better(selected_test) else max(nums)
                    labels.append(d.strftime("%Y-%m-%d"))
                    values.append(float(best_val))
                    summary.append({
                        "date": d,
                        "avg": best_val,
                        "notes": notes,
                    })

        ctx.update({
            "boxer": boxer,
            "tests": tests,
            "selected_test": selected_test,
            "unit": (selected_test.unit if selected_test else ""),
            "labels": labels,
            "values": values,
            "summary": summary,
        })
        return ctx



class BulkBoxerCreateView(LoginRequiredMixin, View):
    template_name = "boxers/bulk_add.html"

    def get(self, request):
        FormSet = formset_factory(BulkBoxerForm, extra=8, can_delete=True)
        formset = FormSet()
        return render(request, self.template_name, {"formset": formset})

    def post(self, request):
        FormSet = formset_factory(BulkBoxerForm, extra=0, can_delete=True)
        formset = FormSet(request.POST)

        gym = user_gym(request)
        if not formset.is_valid():
            return render(request, self.template_name, {"formset": formset})

        existing_qs = Boxer.objects.filter(gym=gym)

        seen_keys = set()
        had_errors = False
        rows_to_create = []

        for form in formset:
            cd = form.cleaned_data or {}
            if cd.get("DELETE"):
                continue

            fn = (cd.get("first_name") or "").strip()
            ln = (cd.get("last_name") or "").strip()
            parent = (cd.get("parent_name") or "").strip()
            dob = cd.get("date_of_birth")

            # Skip truly empty rows
            if not any([fn, ln, parent, dob]):
                continue

            # Require first name
            if not fn:
                form.add_error("first_name", "First name is required.")
                had_errors = True
                continue

            # Build full_name for .name field
            full_name = (fn + " " + ln).strip()

            # Ambiguity & duplicate checks (same as before)
            if not ln and not parent and not dob:
                ambiguous = existing_qs.filter(Q(name__iexact=fn) | Q(name__istartswith=fn + " ")).exists()
                if ambiguous:
                    form.add_error("last_name", "Add last name, parent or birth date to differentiate from existing boxers.")
                    had_errors = True
                    continue

            if ln:
                same_name_qs = existing_qs.filter(name__iexact=full_name)
                if same_name_qs.exists():
                    exact_parent = parent and same_name_qs.filter(parent_name__iexact=parent).exists()
                    exact_dob = dob and same_name_qs.filter(date_of_birth=dob).exists()

                    if (parent and exact_parent) or (dob and exact_dob):
                        form.add_error(None, f"A boxer named “{full_name}” with these details already exists.")
                        had_errors = True
                        continue

                    if not parent and not dob:
                        form.add_error("parent_name", "Add parent or birth date to differentiate from an existing boxer with the same name.")
                        had_errors = True
                        continue

            key = (fn.lower(), ln.lower(), parent.lower(), dob.isoformat() if dob else None)
            if key in seen_keys:
                form.add_error(None, "Duplicate of another row in this submission.")
                had_errors = True
                continue
            seen_keys.add(key)

            # ✅ Save actual first_name / last_name
            rows_to_create.append((fn, ln, full_name, parent, dob))

        if had_errors:
            return render(request, self.template_name, {"formset": formset})

        created = 0
        for fn, ln, full_name, parent, dob in rows_to_create:
            Boxer.objects.create(
                first_name=fn,
                last_name=ln,
                name=full_name,  # keep for backward compatibility
                parent_name=parent or "",
                date_of_birth=dob,
                gym=gym,
            )
            created += 1

        messages.success(
            request,
            f"{created} boxer{'s were' if created != 1 else ' was'} added to {gym}."
        )
        return redirect("mark_attendance")


class BoxerCommentsView(LoginRequiredMixin, View):
    template_name = "boxers/boxer_comments.html"

    def get(self, request, boxer_id):
        boxer = get_object_or_404(Boxer, pk=boxer_id, gym=user_gym(request))
        comments = boxer.comments.select_related("coach").all()
        form = BoxerCommentForm()
        return render(request, self.template_name, {
            "boxer": boxer,
            "comments": comments,
            "form": form,
        })

    def post(self, request, boxer_id):
        boxer = get_object_or_404(Boxer, pk=boxer_id, gym=user_gym(request))
        form = BoxerCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.boxer = boxer
            comment.coach = request.user
            comment.save()
            messages.success(request, "Comment added.")
            return redirect("boxer_comments", boxer_id=boxer.id)

        comments = boxer.comments.select_related("coach").all()
        return render(request, self.template_name, {
            "boxer": boxer,
            "comments": comments,
            "form": form,
        })

class EditCommentView(LoginRequiredMixin, View):
    template_name = "boxers/edit_comment.html"

    def get(self, request, boxer_id, comment_id):
        boxer = get_object_or_404(Boxer, pk=boxer_id, gym=user_gym(request))
        comment = get_object_or_404(BoxerComment, pk=comment_id, boxer=boxer, coach=request.user)
        form = BoxerCommentForm(instance=comment)
        return render(request, self.template_name, {"form": form, "boxer": boxer, "comment": comment})

    def post(self, request, boxer_id, comment_id):
        boxer = get_object_or_404(Boxer, pk=boxer_id, gym=user_gym(request))
        comment = get_object_or_404(BoxerComment, pk=comment_id, boxer=boxer, coach=request.user)
        form = BoxerCommentForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            messages.success(request, "Comment updated.")
            return redirect("boxer_comments", boxer_id=boxer.id)
        return render(request, self.template_name, {"form": form, "boxer": boxer, "comment": comment})


class DeleteCommentView(LoginRequiredMixin, View):
    def post(self, request, boxer_id, comment_id):
        boxer = get_object_or_404(Boxer, pk=boxer_id, gym=user_gym(request))
        comment = get_object_or_404(BoxerComment, pk=comment_id, boxer=boxer, coach=request.user)
        comment.delete()
        messages.success(request, "Comment deleted.")
        return redirect("boxer_comments", boxer_id=boxer.id)


@login_required
def export_attendance_view(request):
    # If the user is a superuser/admin → show all classes
    if request.user.is_superuser or request.user.is_staff:
        classes = ClassTemplate.objects.all()
    else:
        # Otherwise (coach) → filter by their gym
        coach_profile = getattr(request.user, "coach_profile", None)
        if coach_profile and coach_profile.gym:
            classes = ClassTemplate.objects.filter(gym=coach_profile.gym)
        else:
            classes = ClassTemplate.objects.none()

    return render(request, "attendance/export_attendance.html", {"classes": classes})



@login_required
def export_attendance_excel(request):
    class_id = request.GET.get("class_id")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if not (class_id and start_date and end_date):
        return HttpResponse("Missing parameters", status=400)

    try:
        gym_class = ClassTemplate.objects.get(id=class_id)
    except ClassTemplate.DoesNotExist:
        return HttpResponse("Class not found", status=404)

    # Permission: admins/staff ok; coaches only their gym
    if not (request.user.is_superuser or request.user.is_staff):
        coach_profile = getattr(request.user, "coach_profile", None)
        if not coach_profile or coach_profile.gym_id != gym_class.gym_id:
            return HttpResponse("You do not have permission to export this class", status=403)

    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date   = datetime.strptime(end_date,   "%Y-%m-%d").date()

    # 1) Enrolled boxers (the list you asked for)
    enrollments = Enrollment.objects.filter(template=gym_class).select_related("boxer")
    boxers = [en.boxer for en in enrollments]

    # 2) Session dates = any date within range where this class has at least one attendance row
    session_dates = set(
        Attendance.objects.filter(
            class_template=gym_class,
            date__range=(start_date, end_date)
        ).values_list("date", flat=True).distinct()
    )

    # 3) Excel setup
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Export"
    ws.append(["Boxer", "Presents", "Excused Absences", "Unexcused Absences"])

    # 4) Count per boxer
    for boxer in boxers:
        # all rows for this boxer/class/range
        rows = list(
            Attendance.objects.filter(
                boxer=boxer,
                class_template=gym_class,
                date__range=(start_date, end_date),
            ).values("date", "is_present", "is_excused")
        )

        present  = sum(1 for r in rows if r["is_present"])
        excused  = sum(1 for r in rows if (not r["is_present"] and r["is_excused"]))
        explicit_unexcused = sum(1 for r in rows if (not r["is_present"] and not r["is_excused"]))

        # dates this boxer has *any* row
        boxer_dates = {r["date"] for r in rows}
        # treat missing rows on session dates as unexcused
        present = sum(1 for r in rows if r["is_present"])
        excused = sum(1 for r in rows if (not r["is_present"] and r["is_excused"]))
        unexcused = sum(1 for r in rows if (not r["is_present"] and not r["is_excused"]))

        ws.append([str(boxer), present, excused, unexcused])

    # Autosize columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_classname = gym_class.title.replace(" ", "_")
    filename = f"attendance_{safe_classname}_{start_date}_to_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_attendance_preview(request):
    class_id = request.GET.get("class_id")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if not (class_id and start_date and end_date):
        return JsonResponse([], safe=False)

    try:
        gym_class = ClassTemplate.objects.get(id=class_id)
    except ClassTemplate.DoesNotExist:
        return JsonResponse([], safe=False)

    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    enrollments = Enrollment.objects.filter(template=gym_class).select_related("boxer")
    boxers = [en.boxer for en in enrollments]

    result = []
    for boxer in boxers:
        rows = Attendance.objects.filter(
            boxer=boxer,
            class_template=gym_class,
            date__range=(start_date, end_date),
        )
        present = rows.filter(is_present=True).count()
        excused = rows.filter(is_present=False, is_excused=True).count()
        unexcused = rows.filter(is_present=False, is_excused=False).count()
        result.append({
            "boxer": str(boxer),
            "present": present,
            "excused": excused,
            "unexcused": unexcused,
        })

    return JsonResponse(result, safe=False)